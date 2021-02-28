# Example call:
# python3 attendance.py zoomus_meeting_reportXXXXXXX.csv Attendance.xlsx "102(TTH)" "YYYY-MM-DD"

from openpyxl import load_workbook
import csv
import sys

def updateGrades(attendanceFile, sectionNo, nameList, date) -> None:
    wb = load_workbook(filename = attendanceFile)
    mylist = []

    # Select appropriate sheet corresponding to sectionNo's section
    ws = wb[sectionNo]

    numClasses = 50 # Number of class periods in a semester
    nthColumn = 2
    # Finds the right column based on the date we are taking attendance for
    for j in range(2, numClasses):
        if(str(ws.cell(row=1, column=j).value) == date):
            #print("Found: ", j)
            nthColumn = j

    #print(nameList)
    #print("------------------------")
    i = 2
    # Loop over each student name in the first column
    # max_row represents how many students in the section
    #    (it can be > num students, too, but not < num students)
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=40):
        # Only one cell per row due to the max_col constraint
        # All we want to do is loop through the name of each
        #    student in the Attendance Sheet
        for cell in row:
            # If the student's name appears in the Zoom report
            if(cell.value in nameList):
                # Set their attendance to 1 for the respective date
                #    as calculated in nthColumn
                ws.cell(row=i, column=nthColumn).value = 1
            #print(cell.value)
        i += 1
    wb.save(attendanceFile)

def readZoomReport(filename) -> list[str]:
    # Read in list of student names who have attended the meeting
    # Maybe add: Find cumulative meeting time and add a requirement
    #    that total time attending meeting >50% of class time?
    nameList = []
    inputF = open(filename, "r")

    # Read each line of Zoom report
    for line in inputF.readlines():
        tempStr = line.rstrip("\n")
        tempList = tempStr.split(",")

        # Isolate the names
        name = tempList[0]
        nameList.append(name)
    inputF.close()

    # Pop the first record, as it is not a student name
    nameList.pop(0)

    return nameList

# Adds time info (defaults to 00:00:00) to date
def getDateTime(YMD) -> str:
    return (YMD + " 00:00:00")

def main() -> None:
    # .csv file of all attendees of Zoom meeting
    zoomReport = sys.argv[1]

    # .xlsx file of attendance records
    attendanceFile = sys.argv[2]

    # Used to isolate the sheet used for specific section
    sectionNo = sys.argv[3]

    # Specify what date the attendance is for in YYYY-MM-DD format
    dateTime = getDateTime(sys.argv[4])

    nameList = readZoomReport(zoomReport)

    updateGrades(attendanceFile, sectionNo, nameList, dateTime)
    

if __name__ == "__main__":
    main()
